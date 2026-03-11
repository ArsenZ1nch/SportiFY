import sys, os
sys.path.append(os.getcwd())
from configuration.config import Configuration, default_config
from database.database_manager import *
import openpyxl


'''
TODO: make possible to choose sheet
'''
def open_excel_sheet(excel_path: str) -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(excel_path)
    return workbook.active


def get_student_info(worksheet: openpyxl.Workbook) -> dict:
    # map of categories with the corresponding cells
    categorycell_map = {}

    # find cell location of categories
    checks_done = False  # stop checking cell names if all categories found
    for row_tuple in worksheet.rows:
        if not checks_done:
            for cell in row_tuple:
                match cell.value:
                    # TODO: dict keys aus config.json nehmen
                    case "Vorname":
                        categorycell_map["VorName"] = cell
                    case "Nachname":
                        categorycell_map["NachName"] = cell
                    case "Semester":
                        categorycell_map["Semester"] = cell
                    case "1. Wunsch":
                        categorycell_map["Wunsch1"] = cell
                    case "2. Wunsch":
                        categorycell_map["Wunsch2"] = cell
                    case "3. Wunsch":
                        categorycell_map["Wunsch3"] = cell
                    case "4. Wunsch":
                        categorycell_map["Wunsch4"] = cell
                    case "5. Wunsch":
                        categorycell_map["Wunsch5"] = cell
                    case "6. Wunsch":
                        categorycell_map["Wunsch6"] = cell
                        # final category to find -> no more checks needed
                        checks_done = True
        # find last row with data
        last_data_row_tuple = row_tuple

    # expected row of the categories. VorName as example
    categories_row = categorycell_map["VorName"].row

    EXPECTED_CATEGORY_AMOUNT = 9
    if len(categorycell_map) < EXPECTED_CATEGORY_AMOUNT:  # not all categories found
        raise Exception(f"not all categories found. found {len(categorycell_map)}, while {EXPECTED_CATEGORY_AMOUNT} expected")
    if len(categorycell_map) > EXPECTED_CATEGORY_AMOUNT:  # too many categories
        raise Exception(f"too many categories found. found {len(categorycell_map)}, while {EXPECTED_CATEGORY_AMOUNT} expected")

    for category, cell in categorycell_map.items():
        if cell.row != categories_row:  # compare row of category to expected row idx
            raise Exception(f"categories are not in the same row. Anomaly at '{category}' row")

    # determine range of data
    first_data_row = categories_row + 1  # one below categories row
    last_data_row = last_data_row_tuple[0].row  # get row idx of random cell in the row

    # assign an array of values to each category/column
    column_values_map = dict()
    for category, cell in categorycell_map.items():
        column = cell.column
        column_values = list()
        for row_idx in range(first_data_row, last_data_row+1):
            cell = worksheet.cell(column=column, row=row_idx)
            column_values.append(cell.value)
        column_values_map[category] = column_values

    return column_values_map


def get_sportcourse_info(worksheet: openpyxl.Workbook) -> dict:
    # map of categories with the corresponding cells
    categorycell_map = {}

    # find cell location of categories
    checks_done = False  # stop checking cell names if all categories found
    for row_tuple in worksheet.rows:
        if not checks_done:
            for cell in row_tuple:
                match cell.value:
                    # TODO: dict keys aus config.json nehmen
                    case "Kursname":
                        categorycell_map["KursName"] = cell
                    case "Sporthalle":
                        categorycell_map["Sporthalle"] = cell
                    case "Lehrkraft":
                        categorycell_map["Lehrkraft"] = cell
                    case "Themenfeld":
                        categorycell_map["Themenfeld"] = cell
                    case "Plätzeanzahl":
                        categorycell_map["PlatzanzahlMAX"] = cell
                        # final category to find -> no more checks needed
                        checks_done = True
        # find last row with data
        last_data_row_tuple = row_tuple

    # expected row of the categories. VorName as example
    categories_row = categorycell_map["KursName"].row

    EXPECTED_CATEGORY_AMOUNT = 5
    if len(categorycell_map) < EXPECTED_CATEGORY_AMOUNT:  # not all categories found
        raise Exception(f"not all categories found. found {len(categorycell_map)}, while {EXPECTED_CATEGORY_AMOUNT} expected")
    if len(categorycell_map) > EXPECTED_CATEGORY_AMOUNT:  # too many categories
        raise Exception(f"too many categories found. found {len(categorycell_map)}, while {EXPECTED_CATEGORY_AMOUNT} expected")

    for category, cell in categorycell_map.items():
        if cell.row != categories_row:  # compare row of category to expected row idx
            raise Exception(f"categories are not in the same row. Anomaly at '{category}' row")

    # determine range of data
    first_data_row = categories_row + 1  # one below categories row
    last_data_row = last_data_row_tuple[0].row  # get row idx of random cell in the row

    # assign an array of values to each category/column
    column_values_map = dict()
    for category, cell in categorycell_map.items():
        column = cell.column
        column_values = list()
        for row_idx in range(first_data_row, last_data_row+1):
            cell = worksheet.cell(column=column, row=row_idx)
            column_values.append(cell.value)
        column_values_map[category] = column_values

    return column_values_map


# TODO: sonderkurs



if __name__ == "__main__":
    db = DataBase("test/test.db")
    worksheet = open_excel_sheet("test/test.xlsx")
    student_info = get_student_info(worksheet)
